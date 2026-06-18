import time
import argparse
import requests
import pandas as pd
from pathlib import Path
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# ============================================================
# SETTINGS
# ============================================================

BASE_URL = "https://api-dbw.stat.gov.pl/api"

OUT = Path("gus_dbw_trade")
OUT.mkdir(exist_ok=True)

MASTER_FILE = OUT / "gus_import_saldo_PLN_2018_plus_MASTER.xlsx"

# ============================================================
# WKLEJ KLUCZ API TUTAJ
# ============================================================

API_KEY = "rd0rSweA0HdXUfrNpJB5U6vypciTFcvBzuM8kRFarVU="

HEADERS = {"accept": "application/json"}
if API_KEY:
    HEADERS["X-ClientId"] = API_KEY


# ============================================================
# DBW TABLE SETTINGS
# ============================================================

# Przekrój: Polska; Kraje towary; CN
ID_PRZEKROJ = 1136

DATASETS = [
    {
        "sheet": "import_wysylka_PLN",
        "id_zmienna": 222,
        "description": "Import towarów wg kraju wysyłki",
    },
    {
        "sheet": "saldo_pochodzenie_PLN",
        "id_zmienna": 223,
        "description": "Saldo obrotów towarowych wg kraju pochodzenia",
    },
]

# 282 = rok
# 247-258 = miesiące M01-M12
PERIODS = {
    282: {"month": None, "period_type": "year"},

    247: {"month": 1, "period_type": "month"},
    248: {"month": 2, "period_type": "month"},
    249: {"month": 3, "period_type": "month"},
    250: {"month": 4, "period_type": "month"},
    251: {"month": 5, "period_type": "month"},
    252: {"month": 6, "period_type": "month"},
    253: {"month": 7, "period_type": "month"},
    254: {"month": 8, "period_type": "month"},
    255: {"month": 9, "period_type": "month"},
    256: {"month": 10, "period_type": "month"},
    257: {"month": 11, "period_type": "month"},
    258: {"month": 12, "period_type": "month"},
}

# Dokładny wiersz z tabeli DBW:
# POLSKA / Kraje towary = Ogółem / CN = OG306966 - OGÓŁEM
POS_POLSKA = 33617
POS_KRAJE_OGOL = 6649664
POS_CN_OGOL = 7438267

PAGE_SIZE = 5000
MAX_PAGE = 80

# Krótszy sleep, bo masz klucz API
SLEEP = 0.2 if API_KEY else 0.8

# Jeżeli kiedyś debug pokaże stały ID dla [zł], możesz go wpisać tutaj.
# Na razie zostaw None.
FORCE_PLN_WAY_ID = None
DISCOVERED_PLN_WAY_ID = None

# Kontrola tylko dla danych, które porównywałeś ze screenem
EXPECTED_CHECKS = {
    ("Import towarów wg kraju wysyłki", "2018 M01"): 76630519831,
    ("Import towarów wg kraju wysyłki", "2018 M02"): 74282490308,
    ("Import towarów wg kraju wysyłki", "2018 M03"): 82427226760,

    ("Saldo obrotów towarowych wg kraju pochodzenia", "2018 M01"): -2145823146,
    ("Saldo obrotów towarowych wg kraju pochodzenia", "2018 M02"): -2400474546,
    ("Saldo obrotów towarowych wg kraju pochodzenia", "2018 M03"): -1696734268,
}


# ============================================================
# SESSION
# ============================================================

session = requests.Session()

retries = Retry(
    total=5,
    backoff_factor=2,
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["GET"],
)

adapter = HTTPAdapter(max_retries=retries)
session.mount("https://", adapter)
session.mount("http://", adapter)


# ============================================================
# HELPERS
# ============================================================

def get_json(endpoint, params=None, allow_404=False):
    url = f"{BASE_URL}/{endpoint}"

    for attempt in range(1, 6):
        try:
            print(f"REQUEST attempt {attempt}/5: {url}", flush=True)
            print("PARAMS:", params, flush=True)

            r = session.get(
                url,
                params=params,
                headers=HEADERS,
                timeout=30,
            )

            time.sleep(SLEEP)

            print("STATUS:", r.status_code, flush=True)

            if r.status_code == 404 and allow_404:
                return None

            if r.status_code != 200:
                print("FULL URL:", r.url, flush=True)
                print("TEXT:", r.text[:1500], flush=True)
                r.raise_for_status()

            return r.json()

        except requests.exceptions.RequestException as e:
            print(f"Request failed, attempt {attempt}/5: {e}", flush=True)
            time.sleep(5 * attempt)

    raise RuntimeError("API request failed after 5 attempts")


def extract_rows(obj):
    if obj is None:
        return []

    if isinstance(obj, list):
        return obj

    candidates = []

    def walk(x):
        if isinstance(x, list):
            if x and all(isinstance(i, dict) for i in x):
                candidates.append(x)
            for i in x:
                walk(i)
        elif isinstance(x, dict):
            for v in x.values():
                walk(v)

    walk(obj)

    return max(candidates, key=len) if candidates else []


def to_int(x):
    try:
        return int(x)
    except Exception:
        try:
            return int(float(x))
        except Exception:
            return None


def to_float(x):
    try:
        return float(x)
    except Exception:
        return None


def norm(x):
    return str(x or "").strip().lower()


def make_period_label(year, month):
    if month is None:
        return str(year)
    return f"{year} M{month:02d}"


def period_sort_key(col):
    col = str(col)

    if col.isdigit():
        return int(col), 0

    try:
        year, month = col.split(" M")
        return int(year), int(month)
    except Exception:
        return 9999, 99


def get_spm_id(row):
    for k, v in row.items():
        key = str(k).lower().replace("_", "-")
        if "id-sposob-prezentacji-miara" in key:
            return to_int(v)
    return None


def is_target_position(row):
    return (
        to_int(row.get("id-pozycja-1")) == POS_POLSKA
        and to_int(row.get("id-pozycja-2")) == POS_KRAJE_OGOL
        and to_int(row.get("id-pozycja-3")) == POS_CN_OGOL
    )


def build_api_filter():
    return (
        f"id-pozycja-1=={POS_POLSKA},"
        f"id-pozycja-2=={POS_KRAJE_OGOL},"
        f"id-pozycja-3=={POS_CN_OGOL}"
    )


# ============================================================
# API DOWNLOAD
# ============================================================

def fetch_page(id_zmienna, year, id_okres, page, use_api_filter=True):
    params = {
        "id-zmienna": id_zmienna,
        "id-przekroj": ID_PRZEKROJ,
        "id-rok": year,
        "id-okres": id_okres,
        "ile-na-stronie": PAGE_SIZE,
        "numer-strony": page,
        "lang": "pl",
    }

    if use_api_filter:
        params["filters"] = build_api_filter()

    obj = get_json(
        "variable/variable-data-section",
        params=params,
        allow_404=True,
    )

    return extract_rows(obj)


def collect_candidates(dataset, year, id_okres, period_label):
    """
    Najpierw próbuje z filtrem API.
    Jeśli nie znajdzie, robi fallback bez filtra i filtruje lokalnie.
    """
    all_candidates = []

    print("\n--- TRY API FILTER ---", flush=True)

    for page in range(0, 10):
        rows = fetch_page(
            id_zmienna=dataset["id_zmienna"],
            year=year,
            id_okres=id_okres,
            page=page,
            use_api_filter=True,
        )

        if not rows:
            print("filtered page empty, stopping", flush=True)
            break

        matched = [row for row in rows if is_target_position(row)]

        print(
            "filtered page:",
            page,
            "rows:",
            len(rows),
            "matched:",
            len(matched),
            flush=True,
        )

        all_candidates.extend(matched)

        if len(rows) < PAGE_SIZE:
            break

    if all_candidates:
        return all_candidates

    print("\n--- FALLBACK FULL SCAN ---", flush=True)

    for page in range(MAX_PAGE + 1):
        rows = fetch_page(
            id_zmienna=dataset["id_zmienna"],
            year=year,
            id_okres=id_okres,
            page=page,
            use_api_filter=False,
        )

        if not rows:
            print("full scan page empty, stopping", flush=True)
            break

        matched = [row for row in rows if is_target_position(row)]

        print(
            "full page:",
            page,
            "rows:",
            len(rows),
            "matched:",
            len(matched),
            flush=True,
        )

        all_candidates.extend(matched)

        if len(rows) < PAGE_SIZE:
            break

    return all_candidates


def choose_pln_candidate(candidates):
    """
    Dla dokładnego wiersza POLSKA / Ogółem / OG306966 API zwraca kilka typów informacji:
    np. wskaźnik, EUR, USD, zł. Dla tej tabeli wartość [zł] jest największa bezwzględnie.
    Jeżeli wcześniej odkryliśmy ID typu [zł], używamy go w pierwszej kolejności.
    """
    global DISCOVERED_PLN_WAY_ID

    if not candidates:
        raise RuntimeError("Brak kandydatów do wyboru wartości [zł].")

    preferred_way_id = FORCE_PLN_WAY_ID or DISCOVERED_PLN_WAY_ID

    if preferred_way_id is not None:
        by_way = [
            row for row in candidates
            if get_spm_id(row) == preferred_way_id
        ]

        if len(by_way) == 1:
            return by_way[0]

    numeric = []

    for row in candidates:
        value = to_float(row.get("wartosc"))

        if value is None:
            continue

        # Odrzucamy małe wskaźniki typu 96.8 / 105.1
        if abs(value) < 10000:
            continue

        numeric.append(row)

    if not numeric:
        raise RuntimeError("Nie ma liczbowych kandydatów po odrzuceniu wskaźników.")

    selected = max(
        numeric,
        key=lambda r: abs(to_float(r.get("wartosc"))),
    )

    DISCOVERED_PLN_WAY_ID = get_spm_id(selected)

    print("DISCOVERED_PLN_WAY_ID:", DISCOVERED_PLN_WAY_ID, flush=True)

    return selected


def download_one_period(dataset, year, id_okres, period_info):
    month = period_info["month"]
    period_label = make_period_label(year, month)

    print("\n============================================================", flush=True)
    print("DOWNLOADING:", dataset["description"], period_label, flush=True)
    print("id_zmienna:", dataset["id_zmienna"], "id_okres:", id_okres, flush=True)
    print("============================================================", flush=True)

    candidates = collect_candidates(
        dataset=dataset,
        year=year,
        id_okres=id_okres,
        period_label=period_label,
    )

    if not candidates:
        raise RuntimeError(
            f"Nie znaleziono kandydatów dla {dataset['description']} {period_label}"
        )

    debug_df = pd.DataFrame([
        {
            "wartosc": r.get("wartosc"),
            "id_sposob_prezentacji_miara": get_spm_id(r),
            "id-pozycja-1": r.get("id-pozycja-1"),
            "id-pozycja-2": r.get("id-pozycja-2"),
            "id-pozycja-3": r.get("id-pozycja-3"),
        }
        for r in candidates
    ])

    print("\nCANDIDATES:", flush=True)
    print(debug_df.to_string(index=False), flush=True)

    selected = choose_pln_candidate(candidates)
    value = to_float(selected.get("wartosc"))

    key = (dataset["description"], period_label)
    if key in EXPECTED_CHECKS:
        expected = EXPECTED_CHECKS[key]

        if int(value) != int(expected):
            debug_path = OUT / f"debug_BAD_{dataset['sheet']}_{period_label.replace(' ', '_')}.xlsx"
            debug_df.to_excel(debug_path, index=False)

            raise RuntimeError(
                f"NIEZGODNOŚĆ względem DBW dla {dataset['description']} {period_label}: "
                f"jest {value}, powinno być {expected}. "
                f"Debug zapisany: {debug_path}"
            )

    print("SELECTED VALUE:", value, flush=True)

    return {
        "dataset": dataset["sheet"],
        "Zmienna": dataset["description"],
        "id_zmienna": dataset["id_zmienna"],
        "id_przekroj": ID_PRZEKROJ,
        "Typ informacji": "[zł]",
        "Kraje towary": "Ogółem",
        "CN": "OG306966 - OGÓŁEM",
        "Jednostka terytorialna": "POLSKA",
        "year": year,
        "month": month,
        "period": period_label,
        "period_type": period_info["period_type"],
        "wartosc": value,
        "id_sposob_prezentacji_miara": get_spm_id(selected),
    }


def download_dataset_for_year(dataset, year):
    rows = []

    for id_okres, period_info in PERIODS.items():
        row = download_one_period(
            dataset=dataset,
            year=year,
            id_okres=id_okres,
            period_info=period_info,
        )
        rows.append(row)

    return pd.DataFrame(rows)


# ============================================================
# LAYOUT + MASTER FILE
# ============================================================

KEY_COLUMNS = [
    "Zmienna",
    "Typ informacji",
    "Kraje towary",
    "CN",
    "Jednostka terytorialna",
]


def make_gus_layout(df):
    if df.empty:
        return pd.DataFrame()

    df = df.copy()
    df["wartosc"] = pd.to_numeric(df["wartosc"], errors="coerce")

    layout = df.pivot_table(
        index=KEY_COLUMNS,
        columns="period",
        values="wartosc",
        aggfunc="first",
    ).reset_index()

    period_cols = sorted(
        [
            c for c in layout.columns
            if isinstance(c, str) and (c.isdigit() or " M" in c)
        ],
        key=period_sort_key,
    )

    return layout[KEY_COLUMNS + period_cols]


def read_existing_sheet(path, sheet_name):
    if not path.exists():
        return None

    try:
        return pd.read_excel(path, sheet_name=sheet_name)
    except ValueError:
        return None


def combine_existing_with_new(existing_df, new_df, year):
    if existing_df is None or existing_df.empty:
        combined = new_df.copy()
    else:
        existing_df = existing_df.copy()
        new_df = new_df.copy()

        year_str = str(year)
        year_month_prefix = f"{year} M"

        cols_to_drop = [
            c for c in existing_df.columns
            if isinstance(c, str)
            and (c == year_str or c.startswith(year_month_prefix))
        ]

        existing_df = existing_df.drop(columns=cols_to_drop, errors="ignore")

        combined = existing_df.merge(
            new_df,
            on=KEY_COLUMNS,
            how="outer",
        )

    period_cols = [
        c for c in combined.columns
        if isinstance(c, str) and (c.isdigit() or " M" in c)
    ]

    period_cols = sorted(period_cols, key=period_sort_key)

    other_cols = [
        c for c in combined.columns
        if c not in KEY_COLUMNS and c not in period_cols
    ]

    return combined[KEY_COLUMNS + other_cols + period_cols]


def save_master_excel(sheets, summary_df):
    with pd.ExcelWriter(MASTER_FILE, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)

        for sheet_name, df in sheets.items():
            if not df.empty:
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)


def autofit_excel(path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="EDE7F6")
    header_font = Font(bold=True, color="5E35B1")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter

            for cell in col:
                value = cell.value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))

            width = min(max(max_len + 2, 10), 45)
            ws.column_dimensions[col_letter].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

    wb.save(path)


def update_master_excel(year, create_new=False):
    final_sheets = {}
    summary_rows = []

    for dataset in DATASETS:
        print("\n==============================", flush=True)
        print("DATASET:", dataset["sheet"], flush=True)
        print("YEAR:", year, flush=True)
        print("==============================", flush=True)

        raw_df = download_dataset_for_year(
            dataset=dataset,
            year=year,
        )

        raw_csv_path = OUT / f"{dataset['sheet']}_{year}_raw.csv"
        raw_df.to_csv(raw_csv_path, index=False, encoding="utf-8-sig")

        new_layout_df = make_gus_layout(raw_df)

        if create_new:
            existing_df = None
        else:
            existing_df = read_existing_sheet(MASTER_FILE, dataset["sheet"])

        combined_df = combine_existing_with_new(
            existing_df=existing_df,
            new_df=new_layout_df,
            year=year,
        )

        final_sheets[dataset["sheet"]] = combined_df

        summary_rows.append(
            {
                "dataset": dataset["sheet"],
                "description": dataset["description"],
                "updated_year": year,
                "raw_rows_for_year": len(raw_df),
                "final_rows_in_sheet": len(combined_df),
                "api_key_used": bool(API_KEY),
                "raw_csv": str(raw_csv_path),
            }
        )

    summary_df = pd.DataFrame(summary_rows)

    save_master_excel(final_sheets, summary_df)
    autofit_excel(MASTER_FILE)

    print("\nSaved master file:", MASTER_FILE, flush=True)
    print(summary_df.to_string(index=False), flush=True)


# ============================================================
# CLI
# ============================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--create-new", action="store_true")
    args = parser.parse_args()

    print("SCRIPT STARTED", flush=True)
    print("YEAR:", args.year, flush=True)
    print("CREATE_NEW:", args.create_new, flush=True)
    print("API KEY USED:", bool(API_KEY), flush=True)

    update_master_excel(
        year=args.year,
        create_new=args.create_new,
    )
