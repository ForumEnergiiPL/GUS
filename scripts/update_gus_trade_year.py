import time
import argparse
import requests
import pandas as pd
from pathlib import Path
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry


# ============================================================
# SETTINGS
# ============================================================

BASE_URL = "https://api-dbw.stat.gov.pl/api"

OUT = Path("gus_dbw_trade")
OUT.mkdir(exist_ok=True)

MASTER_FILE = OUT / "gus_import_saldo_PLN_2018_plus_MASTER.xlsx"


# ============================================================
# API KEY W KODZIE
# ============================================================

API_KEY = "rd0rSweA0HdXUfrNpJB5U6vypciTFcvBzuM8kRFarVU="

HEADERS = {"accept": "application/json"}
if API_KEY:
    HEADERS["X-ClientId"] = API_KEY


# ============================================================
# DANE DBW
# ============================================================

# Przekrój: Polska; Kraje towary; CN
ID_PRZEKROJ = 1136

DATASETS = [
    {
        "sheet": "import_wysylka_PLN",
        "id_zmienna": 222,
        "description": "Import towarów wg kraju wysyłki",
    },
    {
        "sheet": "saldo_pochodzenie_PLN",
        "id_zmienna": 223,
        "description": "Saldo obrotów towarowych wg kraju pochodzenia",
    },
]

# 282 = rok, 247-258 = miesiące M01-M12
PERIODS = {
    282: {"month": None, "period_type": "year"},

    247: {"month": 1, "period_type": "month"},
    248: {"month": 2, "period_type": "month"},
    249: {"month": 3, "period_type": "month"},
    250: {"month": 4, "period_type": "month"},
    251: {"month": 5, "period_type": "month"},
    252: {"month": 6, "period_type": "month"},
    253: {"month": 7, "period_type": "month"},
    254: {"month": 8, "period_type": "month"},
    255: {"month": 9, "period_type": "month"},
    256: {"month": 10, "period_type": "month"},
    257: {"month": 11, "period_type": "month"},
    258: {"month": 12, "period_type": "month"},
}

# Dokładny wiersz:
# POLSKA / Ogółem / OG306966 - OGÓŁEM
POS_POLSKA = 33617
POS_KRAJE_OGOL = 6649664
POS_CN_OGOL = 7438267

PAGE_SIZE = 5000

# MAX_PAGE = 1 oznacza strony 0 i 1, czyli maksymalnie 2 requesty na okres.
# Nie ustawiaj tu 50/80, bo wtedy GitHub będzie mielił bez sensu.
MAX_PAGE = 1

SLEEP = 0.15 if API_KEY else 0.6


# ============================================================
# SESSION
# ============================================================

session = requests.Session()

retries = Retry(
    total=2,
    backoff_factor=1.5,
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["GET"],
)

adapter = HTTPAdapter(max_retries=retries)
session.mount("https://", adapter)
session.mount("http://", adapter)


# ============================================================
# API HELPERS
# ============================================================

def get_json(endpoint, params=None, allow_404=False):
    url = f"{BASE_URL}/{endpoint}"

    for attempt in range(1, 3):
        try:
            r = session.get(
                url,
                params=params,
                headers=HEADERS,
                timeout=20,
            )

            time.sleep(SLEEP)

            if r.status_code == 404 and allow_404:
                return None

            if r.status_code != 200:
                print("URL:", r.url, flush=True)
                print("STATUS:", r.status_code, flush=True)
                print("TEXT:", r.text[:1000], flush=True)
                r.raise_for_status()

            return r.json()

        except requests.exceptions.RequestException as e:
            print(f"Request failed, attempt {attempt}/2: {e}", flush=True)
            time.sleep(3 * attempt)

    print("API request failed after 2 attempts. Skipping this request.", flush=True)
    return None


def extract_rows(obj):
    if obj is None:
        return []

    if isinstance(obj, list):
        return obj

    candidates = []

    def walk(x):
        if isinstance(x, list):
            if x and all(isinstance(i, dict) for i in x):
                candidates.append(x)
            for i in x:
                walk(i)
        elif isinstance(x, dict):
            for v in x.values():
                walk(v)

    walk(obj)

    return max(candidates, key=len) if candidates else []


def to_int(x):
    try:
        return int(x)
    except Exception:
        try:
            return int(float(x))
        except Exception:
            return None


def to_float(x):
    try:
        return float(x)
    except Exception:
        return None


def make_period_label(year, month):
    if month is None:
        return str(year)
    return f"{year} M{month:02d}"


def period_sort_key(col):
    col = str(col)

    if col.isdigit():
        return int(col), 0

    try:
        year, month = col.split(" M")
        return int(year), int(month)
    except Exception:
        return 9999, 99


def get_spm_id(row):
    for k, v in row.items():
        kk = str(k).lower().replace("_", "-")
        if "id-sposob-prezentacji-miara" in kk:
            return to_int(v)
    return None


# ============================================================
# DOWNLOAD
# ============================================================

def build_filter():
    return (
        f"id-pozycja-1=={POS_POLSKA},"
        f"id-pozycja-2=={POS_KRAJE_OGOL},"
        f"id-pozycja-3=={POS_CN_OGOL}"
    )


def fetch_page(id_zmienna, year, id_okres, page):
    params = {
        "id-zmienna": id_zmienna,
        "id-przekroj": ID_PRZEKROJ,
        "id-rok": year,
        "id-okres": id_okres,
        "filters": build_filter(),
        "ile-na-stronie": PAGE_SIZE,
        "numer-strony": page,
        "lang": "pl",
    }

    obj = get_json(
        "variable/variable-data-section",
        params=params,
        allow_404=True,
    )

    return extract_rows(obj)


def is_target_row(row):
    return (
        to_int(row.get("id-pozycja-1")) == POS_POLSKA
        and to_int(row.get("id-pozycja-2")) == POS_KRAJE_OGOL
        and to_int(row.get("id-pozycja-3")) == POS_CN_OGOL
    )


def choose_pln_candidate(candidates):
    numeric = []

    for row in candidates:
        value = to_float(row.get("wartosc"))

        if value is None:
            continue

        # Odrzucamy wskaźniki typu 96.8, 105.1 itd.
        if abs(value) < 10000:
            continue

        numeric.append(row)

    if not numeric:
        raise RuntimeError("Brak liczbowych kandydatów dla wartości [zł].")

    # Dla tego widoku [zł] jest największe bezwzględnie
    # spośród wartości dla dokładnego wiersza.
    selected = max(
        numeric,
        key=lambda r: abs(to_float(r.get("wartosc"))),
    )

    return selected


def empty_period_row(dataset, year, month, period_label, period_info):
    return {
        "dataset": dataset["sheet"],
        "Zmienna": dataset["description"],
        "id_zmienna": dataset["id_zmienna"],
        "id_przekroj": ID_PRZEKROJ,
        "Typ informacji": "[zł]",
        "Kraje towary": "Ogółem",
        "CN": "OG306966 - OGÓŁEM",
        "Jednostka terytorialna": "POLSKA",
        "year": year,
        "month": month,
        "period": period_label,
        "period_type": period_info["period_type"],
        "wartosc": None,
        "id_sposob_prezentacji_miara": None,
        "status": "missing_or_failed",
    }


def download_one_period(dataset, year, id_okres, period_info):
    month = period_info["month"]
    period_label = make_period_label(year, month)

    print(
        f"\nDownloading: {dataset['sheet']} | {period_label} | id_okres={id_okres}",
        flush=True,
    )

    candidates = []

    for page in range(MAX_PAGE + 1):
        rows = fetch_page(
            id_zmienna=dataset["id_zmienna"],
            year=year,
            id_okres=id_okres,
            page=page,
        )

        if not rows:
            print("page:", page, "empty - stop", flush=True)
            break

        matched = [row for row in rows if is_target_row(row)]

        print(
            "page:", page,
            "rows:", len(rows),
            "matched:", len(matched),
            flush=True,
        )

        candidates.extend(matched)

        # Jeśli znaleźliśmy dokładny wiersz i jest kilka typów informacji,
        # nie mielimy dalej.
        if candidates:
            break

        if len(rows) < PAGE_SIZE:
            break

    if not candidates:
        print(
            f"WARNING: no candidates for {dataset['description']} {period_label}. Saving empty value.",
            flush=True,
        )
        return empty_period_row(dataset, year, month, period_label, period_info)

    try:
        selected = choose_pln_candidate(candidates)
        value = to_float(selected.get("wartosc"))
        spm_id = get_spm_id(selected)
        status = "ok"
    except Exception as e:
        print(
            f"WARNING: cannot choose PLN candidate for {dataset['description']} {period_label}: {e}",
            flush=True,
        )
        value = None
        spm_id = None
        status = "candidate_selection_failed"

    print("selected value:", value, flush=True)

    return {
        "dataset": dataset["sheet"],
        "Zmienna": dataset["description"],
        "id_zmienna": dataset["id_zmienna"],
        "id_przekroj": ID_PRZEKROJ,
        "Typ informacji": "[zł]",
        "Kraje towary": "Ogółem",
        "CN": "OG306966 - OGÓŁEM",
        "Jednostka terytorialna": "POLSKA",
        "year": year,
        "month": month,
        "period": period_label,
        "period_type": period_info["period_type"],
        "wartosc": value,
        "id_sposob_prezentacji_miara": spm_id,
        "status": status,
    }


def download_dataset_for_year(dataset, year):
    rows = []

    for id_okres, period_info in PERIODS.items():
        row = download_one_period(
            dataset=dataset,
            year=year,
            id_okres=id_okres,
            period_info=period_info,
        )
        rows.append(row)

    return pd.DataFrame(rows)


# ============================================================
# LAYOUT
# ============================================================

KEY_COLUMNS = [
    "Zmienna",
    "Typ informacji",
    "Kraje towary",
    "CN",
    "Jednostka terytorialna",
]


def make_gus_layout(df):
    if df.empty:
        return pd.DataFrame()

    df = df.copy()
    df["wartosc"] = pd.to_numeric(df["wartosc"], errors="coerce")

    layout = df.pivot_table(
        index=KEY_COLUMNS,
        columns="period",
        values="wartosc",
        aggfunc="first",
    ).reset_index()

    period_cols = sorted(
        [
            c for c in layout.columns
            if isinstance(c, str) and (c.isdigit() or " M" in c)
        ],
        key=period_sort_key,
    )

    return layout[KEY_COLUMNS + period_cols]


# ============================================================
# MASTER FILE
# ============================================================

def read_existing_sheet(path, sheet_name):
    if not path.exists():
        return None

    try:
        return pd.read_excel(path, sheet_name=sheet_name)
    except ValueError:
        return None


def combine_existing_with_new(existing_df, new_df, year):
    if existing_df is None or existing_df.empty:
        combined = new_df.copy()
    else:
        existing_df = existing_df.copy()
        new_df = new_df.copy()

        year_str = str(year)
        year_month_prefix = f"{year} M"

        # Usuwa stare kolumny danego roku,
        # żeby ponowne odpalenie tego samego roku nadpisało dane.
        cols_to_drop = [
            c for c in existing_df.columns
            if isinstance(c, str)
            and (c == year_str or c.startswith(year_month_prefix))
        ]

        existing_df = existing_df.drop(columns=cols_to_drop, errors="ignore")

        combined = existing_df.merge(
            new_df,
            on=KEY_COLUMNS,
            how="outer",
        )

    period_cols = [
        c for c in combined.columns
        if isinstance(c, str) and (c.isdigit() or " M" in c)
    ]

    period_cols = sorted(period_cols, key=period_sort_key)

    other_cols = [
        c for c in combined.columns
        if c not in KEY_COLUMNS and c not in period_cols
    ]

    return combined[KEY_COLUMNS + other_cols + period_cols]


def save_master_excel(sheets, summary_df):
    with pd.ExcelWriter(MASTER_FILE, engine="openpyxl") as writer:
        summary_df.to_excel(writer, sheet_name="summary", index=False)

        for sheet_name, df in sheets.items():
            if not df.empty:
                df.to_excel(writer, sheet_name=sheet_name[:31], index=False)


def autofit_excel(path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="EDE7F6")
    header_font = Font(bold=True, color="5E35B1")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter

            for cell in col:
                value = cell.value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))

            width = min(max(max_len + 2, 10), 45)
            ws.column_dimensions[col_letter].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

    wb.save(path)


def update_master_excel(year, create_new=False):
    final_sheets = {}
    summary_rows = []

    for dataset in DATASETS:
        print("\n==============================", flush=True)
        print("DATASET:", dataset["sheet"], flush=True)
        print("YEAR:", year, flush=True)
        print("==============================", flush=True)

        raw_df = download_dataset_for_year(dataset, year)

        raw_csv_path = OUT / f"{dataset['sheet']}_{year}_raw.csv"
        raw_df.to_csv(raw_csv_path, index=False, encoding="utf-8-sig")

        new_layout_df = make_gus_layout(raw_df)

        if create_new:
            existing_df = None
        else:
            existing_df = read_existing_sheet(MASTER_FILE, dataset["sheet"])

        combined_df = combine_existing_with_new(
            existing_df=existing_df,
            new_df=new_layout_df,
            year=year,
        )

        final_sheets[dataset["sheet"]] = combined_df

        summary_rows.append(
            {
                "dataset": dataset["sheet"],
                "description": dataset["description"],
                "updated_year": year,
                "raw_rows_for_year": len(raw_df),
                "final_rows_in_sheet": len(combined_df),
                "failed_or_missing_periods": int((raw_df["status"] != "ok").sum()),
                "api_key_used": bool(API_KEY),
                "raw_csv": str(raw_csv_path),
            }
        )

    summary_df = pd.DataFrame(summary_rows)

    save_master_excel(final_sheets, summary_df)
    autofit_excel(MASTER_FILE)

    print("\nSaved master file:", MASTER_FILE, flush=True)
    print(summary_df.to_string(index=False), flush=True)


# ============================================================
# CLI
# ============================================================

if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--year", type=int, required=True)
    parser.add_argument("--create-new", action="store_true")
    args = parser.parse_args()

    print("SCRIPT STARTED", flush=True)
    print("YEAR:", args.year, flush=True)
    print("CREATE_NEW:", args.create_new, flush=True)
    print("API KEY USED:", bool(API_KEY), flush=True)

    update_master_excel(
        year=args.year,
        create_new=args.create_new,
    )
