import os
import requests
import time
import pandas as pd
from pathlib import Path
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

# ============================================================
# SETTINGS
# ============================================================

BASE_URL = "https://api-dbw.stat.gov.pl/api"

OUT = Path("gus_dbw_trade")
OUT.mkdir(exist_ok=True)

OUTPUT_FILE = OUT / "gus_gas_TJ_2018_plus_IMPORT_EXPORT.xlsx"

API_KEY = os.getenv("GUS_DBW_API_KEY")

HEADERS = {"accept": "application/json"}
if API_KEY:
    HEADERS["X-ClientId"] = API_KEY

# Correct section:
# Polska; Kraje towary; CN - uzupełniająca jednostka miary
ID_PRZEKROJ = 1434

# Years from 2018
YEARS = range(2018, 2026)   # 2018-2025; zmień na range(2018, 2027), jeżeli chcesz też 2026

# All months
MONTHS = {
    247: 1,
    248: 2,
    249: 3,
    250: 4,
    251: 5,
    252: 6,
    253: 7,
    254: 8,
    255: 9,
    256: 10,
    257: 11,
    258: 12,
}

TARGET_CODES = [
    "27111100",  # Gaz ziemny skroplony [TJ]
    "27112100",  # Gaz ziemny w stanie gazowym [TJ]
]

# ONLY TWO DATASETS NOW
DATASETS = [
    {
        "sheet": "import_origin_gas_TJ",
        "id_zmienna": 221,
        "description": "Import towarów wg kraju pochodzenia",
    },
    {
        "sheet": "export_gas_TJ",
        "id_zmienna": 220,
        "description": "Eksport towarów",
    },
]

PAGE_SIZE = 5000
MAX_PAGE = 100
SLEEP = 0.7

# True = only "Ogółem", like on the GUS screen
# False = all countries
ONLY_TOTAL_COUNTRY = True

SAVE_RAW_SHEETS = False


# ============================================================
# SESSION
# ============================================================

session = requests.Session()

retries = Retry(
    total=5,
    backoff_factor=2,
    status_forcelist=[429, 500, 502, 503, 504],
    allowed_methods=["GET"],
)

adapter = HTTPAdapter(max_retries=retries)
session.mount("https://", adapter)
session.mount("http://", adapter)


# ============================================================
# API HELPERS
# ============================================================

def get_json(endpoint, params=None, allow_404=False):
    url = f"{BASE_URL}/{endpoint}"

    for attempt in range(1, 6):
        try:
            r = session.get(
                url,
                params=params,
                headers=HEADERS,
                timeout=90,
            )

            time.sleep(SLEEP)

            if r.status_code == 404 and allow_404:
                return None

            if r.status_code != 200:
                print("URL:", r.url)
                print("STATUS:", r.status_code)
                print("TEXT:", r.text[:1000])
                r.raise_for_status()

            return r.json()

        except requests.exceptions.RequestException as e:
            print(f"Request failed, attempt {attempt}/5:", e)
            time.sleep(5 * attempt)

    raise RuntimeError("API request failed after 5 attempts")


def extract_rows(obj):
    if obj is None:
        return None

    if isinstance(obj, list):
        return obj

    candidates = []

    def walk(x):
        if isinstance(x, list):
            if x and all(isinstance(i, dict) for i in x):
                candidates.append(x)
            for i in x:
                walk(i)
        elif isinstance(x, dict):
            for v in x.values():
                walk(v)

    walk(obj)

    return max(candidates, key=len) if candidates else []


def fetch_page(id_zmienna, year, id_okres, page):
    params = {
        "id-zmienna": id_zmienna,
        "id-przekroj": ID_PRZEKROJ,
        "id-rok": year,
        "id-okres": id_okres,
        "ile-na-stronie": PAGE_SIZE,
        "numer-strony": page,
        "lang": "pl",
    }

    obj = get_json(
        "variable/variable-data-section",
        params=params,
        allow_404=True,
    )

    return extract_rows(obj)


# ============================================================
# METADATA
# ============================================================

def load_sections():
    path = OUT / "03_section_positions.xlsx"

    if not path.exists():
        raise FileNotFoundError(
            "Brakuje pliku: gus_dbw_trade/03_section_positions.xlsx"
        )

    return pd.read_excel(path)


def get_target_product_ids(sections):
    products = sections[
        (sections["id-przekroj"] == ID_PRZEKROJ)
        & (sections["symbol"].astype(str).isin(TARGET_CODES))
    ].copy()

    if products.empty:
        raise RuntimeError(
            "Nie znaleziono kodów 27111100 i 27112100 dla przekroju 1434."
        )

    print("\nTarget products:")
    print(
        products[
            [
                "id-przekroj",
                "id-wymiar",
                "nazwa-wymiar",
                "id-pozycja",
                "symbol",
                "nazwa-pozycja",
            ]
        ].to_string(index=False)
    )

    return products["id-pozycja"].astype(int).tolist()


def get_total_country_ids(sections):
    countries = sections[
        (sections["id-przekroj"] == ID_PRZEKROJ)
        & (sections["id-wymiar"] == 734)
        & (sections["nazwa-pozycja"].astype(str).str.lower().eq("ogółem"))
    ].copy()

    return countries["id-pozycja"].astype(int).tolist()


# ============================================================
# DOWNLOAD
# ============================================================

def filter_rows(rows, target_product_ids, total_country_ids):
    if not rows:
        return []

    out = []

    for row in rows:
        if row.get("id-pozycja-3") not in target_product_ids:
            continue

        if ONLY_TOTAL_COUNTRY:
            if row.get("id-pozycja-2") not in total_country_ids:
                continue

        out.append(row)

    return out


def download_dataset(dataset, target_product_ids, total_country_ids):
    all_rows = []

    for year in YEARS:
        for id_okres, month in MONTHS.items():

            print("\nDownloading:")
            print(
                "dataset:", dataset["sheet"],
                "id_zmienna:", dataset["id_zmienna"],
                "year:", year,
                "month:", month,
                "id_okres:", id_okres,
            )

            for page in range(MAX_PAGE + 1):
                rows = fetch_page(
                    id_zmienna=dataset["id_zmienna"],
                    year=year,
                    id_okres=id_okres,
                    page=page,
                )

                if rows is None:
                    print("page:", page, "does not exist - stopping")
                    break

                matched = filter_rows(
                    rows=rows,
                    target_product_ids=target_product_ids,
                    total_country_ids=total_country_ids,
                )

                print(
                    "page:", page,
                    "rows:", len(rows),
                    "matched:", len(matched),
                )

                for row in matched:
                    row["dataset"] = dataset["sheet"]
                    row["dataset_description"] = dataset["description"]
                    row["year"] = year
                    row["month"] = month
                    row["date"] = f"{year}-{month:02d}-01"
                    row["page_downloaded"] = page

                all_rows.extend(matched)

                if len(rows) < PAGE_SIZE:
                    print("last page reached")
                    break

    if not all_rows:
        return pd.DataFrame()

    df = pd.json_normalize(all_rows)
    df = df.drop_duplicates()

    return df


# ============================================================
# NAMES + GUS-LIKE LAYOUT
# ============================================================

def add_names(df, sections):
    if df.empty:
        return df

    # territory dimension: id-wymiar = 2
    territory = sections[
        (sections["id-przekroj"] == ID_PRZEKROJ)
        & (sections["id-wymiar"] == 2)
    ][["id-pozycja", "nazwa-pozycja", "symbol"]].drop_duplicates(subset=["id-pozycja"])

    territory = territory.rename(
        columns={
            "id-pozycja": "id-pozycja-1",
            "nazwa-pozycja": "territory_name",
            "symbol": "territory_symbol",
        }
    )

    # country dimension: id-wymiar = 734
    countries = sections[
        (sections["id-przekroj"] == ID_PRZEKROJ)
        & (sections["id-wymiar"] == 734)
    ][["id-pozycja", "nazwa-pozycja", "symbol"]].drop_duplicates(subset=["id-pozycja"])

    countries = countries.rename(
        columns={
            "id-pozycja": "id-pozycja-2",
            "nazwa-pozycja": "country_name",
            "symbol": "country_symbol",
        }
    )

    # product dimension: id-wymiar = 1181
    products = sections[
        (sections["id-przekroj"] == ID_PRZEKROJ)
        & (sections["id-wymiar"] == 1181)
        & (sections["symbol"].astype(str).isin(TARGET_CODES))
    ][["id-pozycja", "nazwa-pozycja", "symbol"]].drop_duplicates(subset=["id-pozycja"])

    products = products.rename(
        columns={
            "id-pozycja": "id-pozycja-3",
            "nazwa-pozycja": "product_name",
            "symbol": "product_code",
        }
    )

    df = df.merge(territory, on="id-pozycja-1", how="left")
    df = df.merge(countries, on="id-pozycja-2", how="left")
    df = df.merge(products, on="id-pozycja-3", how="left")

    return df


def make_gus_layout(df, sections):
    if df.empty:
        return pd.DataFrame()

    df = add_names(df, sections)

    df["wartosc"] = pd.to_numeric(df["wartosc"], errors="coerce")

    df["period"] = (
        df["year"].astype(int).astype(str)
        + " M"
        + df["month"].astype(int).astype(str).str.zfill(2)
    )

    df["typ_informacji"] = "[-]"

    df["territory_name"] = df["territory_name"].fillna("POLSKA")
    df["country_name"] = df["country_name"].fillna("")
    df["product_name"] = df["product_name"].fillna(df["product_code"].astype(str))

    key_cols = [
        "dataset_description",
        "typ_informacji",
        "country_name",
        "product_name",
        "territory_name",
        "period",
    ]

    df = df.drop_duplicates(subset=key_cols + ["wartosc"])

    layout = df.pivot_table(
        index=[
            "dataset_description",
            "typ_informacji",
            "country_name",
            "product_name",
            "territory_name",
        ],
        columns="period",
        values="wartosc",
        aggfunc="first",
    ).reset_index()

    period_cols = sorted(
        [
            c for c in layout.columns
            if isinstance(c, str) and " M" in c
        ]
    )

    layout = layout[
        [
            "dataset_description",
            "typ_informacji",
            "country_name",
            "product_name",
            "territory_name",
        ]
        + period_cols
    ]

    layout = layout.rename(
        columns={
            "dataset_description": "Zmienna",
            "typ_informacji": "Typ informacji",
            "country_name": "Kraje towary",
            "product_name": "CN - uzupełniająca jednostka miary",
            "territory_name": "Jednostka terytorialna",
        }
    )

    return layout


# ============================================================
# EXCEL FORMATTING
# ============================================================

def autofit_excel(path):
    from openpyxl import load_workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    wb = load_workbook(path)

    header_fill = PatternFill("solid", fgColor="EDE7F6")
    header_font = Font(bold=True, color="5E35B1")
    thin = Side(style="thin", color="DDDDDD")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    for ws in wb.worksheets:
        ws.freeze_panes = "A2"

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = border

        for row in ws.iter_rows():
            for cell in row:
                cell.border = border
                cell.alignment = Alignment(vertical="center", wrap_text=True)

        for col in ws.columns:
            max_len = 0
            col_letter = col[0].column_letter

            for cell in col:
                value = cell.value
                if value is None:
                    continue
                max_len = max(max_len, len(str(value)))

            width = min(max(max_len + 2, 10), 45)
            ws.column_dimensions[col_letter].width = width

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                if isinstance(cell.value, (int, float)):
                    cell.number_format = '#,##0'

    wb.save(path)


# ============================================================
# MAIN
# ============================================================

sections = load_sections()
target_product_ids = get_target_product_ids(sections)
total_country_ids = get_total_country_ids(sections)

print("\nTotal country IDs:")
print(total_country_ids)

summary_rows = []
raw_results = {}
layout_results = {}

for dataset in DATASETS:
    print("\n==============================")
    print("DATASET:", dataset["sheet"])
    print("==============================")

    raw_df = download_dataset(
        dataset=dataset,
        target_product_ids=target_product_ids,
        total_country_ids=total_country_ids,
    )

    layout_df = make_gus_layout(raw_df, sections)

    raw_results[dataset["sheet"]] = raw_df
    layout_results[dataset["sheet"]] = layout_df

    summary_rows.append(
        {
            "dataset": dataset["sheet"],
            "description": dataset["description"],
            "raw_rows": len(raw_df),
            "final_rows": len(layout_df),
            "first_year": min(YEARS),
            "last_year": max(YEARS),
            "only_total_country": ONLY_TOTAL_COUNTRY,
        }
    )

summary_df = pd.DataFrame(summary_rows)

with pd.ExcelWriter(OUTPUT_FILE, engine="openpyxl") as writer:
    summary_df.to_excel(writer, sheet_name="summary", index=False)

    for sheet_name, layout_df in layout_results.items():
        if not layout_df.empty:
            layout_df.to_excel(writer, sheet_name=sheet_name[:31], index=False)

    if SAVE_RAW_SHEETS:
        for sheet_name, raw_df in raw_results.items():
            if not raw_df.empty:
                raw_df.to_excel(writer, sheet_name=("raw_" + sheet_name)[:31], index=False)

autofit_excel(OUTPUT_FILE)

print("\nSaved:", OUTPUT_FILE)
print(summary_df)
